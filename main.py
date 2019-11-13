import web
import json
import person
import shutil
import mparser
import time
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook

def getDict(t):
    with open("ReductionsJson/t" + str(t) + ".json", "r") as read_file:
        return json.load(read_file)

def main():
    #shutil.copyfile(r"ts.xlsx", r"ts1.xlsx")
    time.sleep(1)
    #Кароч метод LoadFile не хочет возвращать wb, вывод питон - х##ня
    name = 'Files/Strelova_Ekaterina_Vasilyevna'
    wb = load_workbook(name + ".xlsx", data_only=True)
    sheet = wb.active

    p = person.MakePersonList(sheet)
    for perid in range(len(p)):
        print(perid)
        persons = []
        info = p[perid]['info']
        p[perid] = mparser.parse(info, p[perid])
        params = {'Фамилия': p[perid]['surname'], 'Имя': p[perid]['name'], 'Отчество': p[perid]['faname'], 'Звание': p[perid]['rank'], 'Дата рождения/Возраст': p[perid]['birthyear'], 'Дата выбытия': p[perid]['deathday'], 'Причина выбытия': p[perid]['reason']}
        persons = web.getRecords(1, params, [])
        max = 0
        maxpersons = []
        print(persons)
        if (persons != []):
            print('======================================')
            for i in range(len(persons)):
                if (persons[i]['percent'] > max):
                    max = persons[i]['percent']
            for i in range(len(persons)):
                if (persons[i]['percent'] == max):
                    maxpersons.append(persons[i])
            if (maxpersons != []):
                max = 0
                maxindex = -1
                print(len(maxpersons))
                for i in range(0, len(maxpersons)):
                    print(len(str(maxpersons[i])) > max)
                    if len(str(maxpersons[i])) > max: 
                        max = len(str(maxpersons[i]))
                        maxindex = i
                print('!!!', maxpersons[maxindex], '!!!')
                print('***************************************')
                p[perid]['link'] = "https://obd-memorial.ru/html/info.htm?id=" + maxpersons[maxindex]['id']
    size = person.GetSize(sheet)
    for i in range(2, size + 1):
        j = 6
        while (j < 22):
            if (p[i-2][person.field_name[j-1]] != None):
                sheet[get_column_letter(j) + str(i)].value = p[i-2][person.field_name[j-1]]
            else:
                sheet[get_column_letter(j) + str(i)].value = ""
            j += 1
    wb.save(name + "final.xlsx")

if __name__ == "__main__":
    main()