import json
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook

def getWb():
    wb = load_workbook('reductions.xlsx')
    for i in range(1,14) :
        sheet = wb['t' + str(i)]
        j = 2
        data = {}
        while (sheet['A' + str(j)].value != None):
            data[sheet['A' + str(j)].value] = sheet['B' + str(j)].value
            print(sheet['A' + str(j)].value, " - ",sheet['B' + str(j)].value ) 
            j += 1
        print(data) 

        with open("t" + str(i) + ".json", "w") as write_file:
           json.dump(data, write_file)

def main():
    getWb()

if __name__ == "__main__":
    main()