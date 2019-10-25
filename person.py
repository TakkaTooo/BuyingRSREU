from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook

#Ключи
field_name = ['id', 'surname', 'name', 'faname', 
              'info', 'birthyear', 'born_place', 'rank',
              'spec', 'duty_place', 'deathday', 'reason',
              'burial_place', 'burial_ro', 'front', 'add',
              'link', 'note', 'area', 'hero',
              'birthday', 'vol']

#Определяет кол-во записей
def GetSize(_sheet):
  i = 1
  while _sheet['A' + str(i)].value != None:
    i += 1
  return i - 1

#Превращает входные параметры словарь
def ConcatLists(_fields, _params):
  dic = {}
  for i in range(len(_fields)):
    dic.update({_fields[i] : _params[i]})
  return dic

#Получает весь список словарей
def MakePersonList(_sheet):
  person_list = []
  size = GetSize(_sheet)
  for i in range(2, size + 1):
    j = 1
    params = []
    while (j < 23):
      if (_sheet[get_column_letter(j) + str(i)].value == None):
        params.append('')
      else:
        params.append(str(_sheet[get_column_letter(j) + str(i)].value))
      j += 1
    pers = ConcatLists(field_name, params)
    person_list.append(pers)
  return person_list

#Возвращает активный лист(первый) в книге
def LoadFile(file_name):
    wb = load_workbook(filename=file_name+".xlsx", data_only=True)
    return wb.active
